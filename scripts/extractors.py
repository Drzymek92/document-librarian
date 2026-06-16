import base64
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

from scripts import config

# A vision function takes (image_path, model_name) and returns transcribed text.
VisionFn = Callable[[str, str], str]

_EXT_TO_TYPE = {
    ".pdf": "pdf",
    ".docx": "docx",
    ".xlsx": "xlsx",
    ".xlsm": "xlsx",
    ".csv": "csv",
    ".tsv": "csv",
    ".json": "json",
    ".jsonl": "json",
    ".html": "html",
    ".htm": "html",
    ".txt": "text",
    ".md": "text",
    ".png": "image",
    ".jpg": "image",
    ".jpeg": "image",
    ".gif": "image",
    ".bmp": "image",
    ".webp": "image",
}

_CSV_ROWS_PER_CHUNK = 50
_MAX_CHARS = 1200


@dataclass
class ExtractResult:
    source_type: str
    chunks: list[dict[str, Any]]
    title: str | None = None
    author: str | None = None
    page_count: int | None = None
    metadata: dict[str, Any] = field(default_factory=dict)


def _chunk_text(text: str, max_chars: int = _MAX_CHARS) -> list[str]:
    paras = [p.strip() for p in text.split("\n\n") if p.strip()]
    chunks: list[str] = []
    buf = ""
    for p in paras:
        if buf and len(buf) + len(p) + 2 > max_chars:
            chunks.append(buf)
            buf = p
        else:
            buf = f"{buf}\n\n{p}" if buf else p
    if buf:
        chunks.append(buf)
    if not chunks and text.strip():
        chunks = [text.strip()]
    return chunks


def _ocr_pdf_page(page, image_path: Path, model: str, vision_fn: VisionFn) -> str:
    # Rasterize the page at 2x and transcribe via the vision model.
    import fitz

    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
    pix.save(str(image_path))
    return (vision_fn(str(image_path), model) or "").strip()


def _extract_pdf(
    path: Path, effort: str | None = None, vision_fn: VisionFn | None = None
) -> ExtractResult:
    import fitz  # PyMuPDF

    profile = config.get_profile(effort)
    ocr_fn = vision_fn or _default_vision_transcribe
    doc = fitz.open(path)
    chunks: list[dict[str, Any]] = []
    ordinal = 0
    ocr_pages = 0
    for page_idx in range(doc.page_count):
        page = doc.load_page(page_idx)
        page_text = page.get_text().strip()
        locator = f"page {page_idx + 1}"
        if not page_text:
            # Image-only page: text layer is empty. OCR it on tiers that allow it.
            if profile.pdf_ocr_fallback:
                tmp = path.with_suffix(f".p{page_idx + 1}.ocr.png")
                try:
                    page_text = _ocr_pdf_page(page, tmp, profile.vision_model, ocr_fn)
                    locator = f"page {page_idx + 1} (ocr)"
                    ocr_pages += 1
                finally:
                    tmp.unlink(missing_ok=True)
            if not page_text:
                continue
        for piece in _chunk_text(page_text):
            chunks.append({"ordinal": ordinal, "text": piece, "locator": locator})
            ordinal += 1
    meta = doc.metadata or {}
    result = ExtractResult(
        source_type="pdf",
        chunks=chunks,
        title=(meta.get("title") or "").strip() or None,
        author=(meta.get("author") or "").strip() or None,
        page_count=doc.page_count,
        metadata={"ocr_pages": ocr_pages} if ocr_pages else {},
    )
    doc.close()
    return result


def _docx_table_text(table) -> str:
    # Render a table as comma-joined rows so cell content is searchable.
    lines = []
    for row in table.rows:
        cells = [c.text.strip().replace("\n", " ") for c in row.cells]
        if any(cells):
            lines.append(",".join(cells))
    return "\n".join(lines)


def _iter_docx_blocks(document):
    # Walk the body in document order so tables are captured inline with
    # paragraphs — python-docx's `document.paragraphs` excludes table cells.
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield "para", Paragraph(child, document)
        elif isinstance(child, CT_Tbl):
            yield "table", Table(child, document)


def _extract_docx(path: Path) -> ExtractResult:
    import docx

    document = docx.Document(str(path))
    blocks: list[str] = []
    first_para: str | None = None
    for kind, block in _iter_docx_blocks(document):
        if kind == "para":
            text = block.text.strip()
            if text:
                blocks.append(text)
                if first_para is None:
                    first_para = text
        else:
            table_text = _docx_table_text(block)
            if table_text:
                blocks.append(table_text)
    full_text = "\n\n".join(blocks)
    chunks = [
        {"ordinal": i, "text": piece, "locator": f"block {i + 1}"}
        for i, piece in enumerate(_chunk_text(full_text))
    ]
    core = document.core_properties
    return ExtractResult(
        source_type="docx",
        chunks=chunks,
        title=(core.title or first_para) or None,
        author=core.author or None,
    )


def _extract_xlsx(path: Path) -> ExtractResult:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    chunks: list[dict[str, Any]] = []
    ordinal = 0
    for ws in wb.worksheets:
        lines = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(cells):
                lines.append(",".join(cells))
        if not lines:
            continue
        # keep header with each chunk for context
        header = lines[0]
        body = lines[1:] or lines
        for start in range(0, len(body), _CSV_ROWS_PER_CHUNK):
            block = body[start:start + _CSV_ROWS_PER_CHUNK]
            text = "\n".join([header, *block]) if lines[1:] else "\n".join(block)
            chunks.append({
                "ordinal": ordinal,
                "text": text,
                "locator": f"sheet '{ws.title}'",
            })
            ordinal += 1
    wb.close()
    return ExtractResult(source_type="xlsx", chunks=chunks)


def _sniff_csv(path: Path) -> tuple[str, str]:
    # Locale exports are often ;-delimited and cp1252-encoded; the default
    # comma/utf-8 assumption collapses them into one column or raises.
    import csv

    raw = path.read_bytes()
    encoding = "utf-8"
    try:
        sample = raw[:8192].decode("utf-8")
    except UnicodeDecodeError:
        encoding = "cp1252"
        sample = raw[:8192].decode("cp1252", errors="replace")
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ","
    return delimiter, encoding


def _extract_csv(path: Path) -> ExtractResult:
    import pandas as pd

    delimiter, encoding = _sniff_csv(path)
    df = pd.read_csv(
        path, dtype=str, keep_default_na=False, sep=delimiter, encoding=encoding
    )
    header = ",".join(df.columns.astype(str))
    chunks: list[dict[str, Any]] = []
    ordinal = 0
    for start in range(0, len(df), _CSV_ROWS_PER_CHUNK):
        block = df.iloc[start:start + _CSV_ROWS_PER_CHUNK]
        rows = ["" if r is None else ",".join(map(str, r)) for r in block.values]
        text = "\n".join([header, *rows])
        end = min(start + _CSV_ROWS_PER_CHUNK, len(df))
        chunks.append({
            "ordinal": ordinal,
            "text": text,
            "locator": f"rows {start + 1}-{end}",
        })
        ordinal += 1
    return ExtractResult(source_type="csv", chunks=chunks, page_count=len(df))


def _flatten_json(obj: Any, prefix: str = "") -> list[str]:
    # Flatten nested structures to "a.b[0].c: value" lines so every leaf value
    # is plain text and BM25-searchable.
    if isinstance(obj, dict):
        lines: list[str] = []
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            lines.extend(_flatten_json(v, key))
        return lines
    if isinstance(obj, list):
        lines = []
        for i, v in enumerate(obj):
            lines.extend(_flatten_json(v, f"{prefix}[{i}]"))
        return lines
    val = "" if obj is None else str(obj)
    return [f"{prefix}: {val}" if prefix else val]


def _load_json_records(path: Path) -> list[Any]:
    raw = path.read_text(encoding="utf-8", errors="replace").strip()
    if not raw:
        return []
    if path.suffix.lower() == ".jsonl":
        return [json.loads(line) for line in raw.splitlines() if line.strip()]
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        # A .json file that's actually line-delimited — fall back to JSONL.
        return [json.loads(line) for line in raw.splitlines() if line.strip()]
    return data if isinstance(data, list) else [data]


def _extract_json(path: Path) -> ExtractResult:
    records = _load_json_records(path)
    chunks: list[dict[str, Any]] = []
    ordinal = 0
    buf: list[tuple[int, str]] = []
    buf_chars = 0

    def flush() -> None:
        nonlocal ordinal, buf, buf_chars
        if not buf:
            return
        first, last = buf[0][0] + 1, buf[-1][0] + 1
        loc = f"record {first}" if first == last else f"records {first}-{last}"
        chunks.append({"ordinal": ordinal, "text": "\n\n".join(b for _, b in buf), "locator": loc})
        ordinal += 1
        buf = []
        buf_chars = 0

    for i, rec in enumerate(records):
        block = "\n".join(_flatten_json(rec))
        if len(block) > _MAX_CHARS:
            flush()
            for piece in _chunk_text(block):
                chunks.append({"ordinal": ordinal, "text": piece, "locator": f"record {i + 1}"})
                ordinal += 1
            continue
        if buf and buf_chars + len(block) > _MAX_CHARS:
            flush()
        buf.append((i, block))
        buf_chars += len(block) + 2
    flush()
    return ExtractResult(source_type="json", chunks=chunks, page_count=len(records))


def _extract_html(path: Path) -> ExtractResult:
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="replace"), "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    title = soup.title.string.strip() if soup.title and soup.title.string else None
    text = re.sub(r"\n{3,}", "\n\n", soup.get_text("\n")).strip()
    chunks = [
        {"ordinal": i, "text": piece, "locator": f"part {i + 1}"}
        for i, piece in enumerate(_chunk_text(text))
    ]
    return ExtractResult(source_type="html", chunks=chunks, title=title)


def _extract_text(path: Path) -> ExtractResult:
    raw = path.read_text(encoding="utf-8", errors="replace")
    chunks = [
        {"ordinal": i, "text": piece, "locator": f"part {i + 1}"}
        for i, piece in enumerate(_chunk_text(raw))
    ]
    return ExtractResult(source_type="text", chunks=chunks)


_MD_HEADING = re.compile(r"^(#{1,6})\s+(.*\S)\s*$")


def _split_markdown_sections(raw: str) -> list[tuple[str, str]]:
    # Split on ATX headings, keeping each section's body under its heading so
    # the heading can serve as the chunk locator. Text before the first
    # heading is grouped under "preamble".
    sections: list[tuple[str, list[str]]] = [("preamble", [])]
    for line in raw.splitlines():
        m = _MD_HEADING.match(line)
        if m:
            sections.append((m.group(2).strip(), []))
        else:
            sections[-1][1].append(line)
    out = []
    for heading, lines in sections:
        body = "\n".join(lines).strip()
        if body or heading != "preamble":
            out.append((heading, body))
    return out


def _extract_markdown(path: Path) -> ExtractResult:
    raw = path.read_text(encoding="utf-8", errors="replace")
    chunks: list[dict[str, Any]] = []
    ordinal = 0
    title: str | None = None
    for heading, body in _split_markdown_sections(raw):
        if title is None and heading != "preamble":
            title = heading
        section_text = body if heading == "preamble" else f"{heading}\n\n{body}".strip()
        for piece in _chunk_text(section_text):
            chunks.append({"ordinal": ordinal, "text": piece, "locator": heading})
            ordinal += 1
    return ExtractResult(source_type="text", chunks=chunks, title=title)


def _default_vision_transcribe(image_path: str, model: str) -> str:
    from langchain_core.messages import HumanMessage

    from scripts.llm_client import get_llm

    data = Path(image_path).read_bytes()
    b64 = base64.b64encode(data).decode()
    ext = Path(image_path).suffix.lstrip(".").lower() or "png"
    mime = "jpeg" if ext in ("jpg", "jpeg") else ext
    llm = get_llm(model=model)
    msg = HumanMessage(content=[
        {
            "type": "text",
            "text": (
                "Transcribe all text in this image exactly. Then, on a new line, "
                "briefly describe any important non-text visual content. "
                "Return plain text only."
            ),
        },
        {"type": "image_url", "image_url": {"url": f"data:image/{mime};base64,{b64}"}},
    ])
    return llm.invoke([msg]).content


def _extract_image(
    path: Path, effort: str | None, vision_fn: VisionFn | None
) -> ExtractResult:
    profile = config.get_profile(effort)
    fn = vision_fn or _default_vision_transcribe
    text = (fn(str(path), profile.vision_model) or "").strip()
    chunks = [
        {"ordinal": i, "text": piece, "locator": "image"}
        for i, piece in enumerate(_chunk_text(text))
    ]
    return ExtractResult(
        source_type="image",
        chunks=chunks,
        metadata={"vision_model": profile.vision_model},
    )


def extract(
    path: str | Path,
    *,
    effort: str | None = None,
    vision_fn: VisionFn | None = None,
) -> ExtractResult:
    path = Path(path)
    source_type = _EXT_TO_TYPE.get(path.suffix.lower())
    if source_type is None:
        raise ValueError(f"Unsupported file type: {path.suffix} ({path.name})")
    if source_type == "pdf":
        return _extract_pdf(path, effort, vision_fn)
    if source_type == "docx":
        return _extract_docx(path)
    if source_type == "xlsx":
        return _extract_xlsx(path)
    if source_type == "csv":
        return _extract_csv(path)
    if source_type == "json":
        return _extract_json(path)
    if source_type == "html":
        return _extract_html(path)
    if source_type == "text":
        if path.suffix.lower() == ".md":
            return _extract_markdown(path)
        return _extract_text(path)
    return _extract_image(path, effort, vision_fn)
